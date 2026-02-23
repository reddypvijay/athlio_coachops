"""
CoachOps — Export Attendance Sheet (tools/export_attendance_sheet.py)

Reads attendance records for a community + month from Supabase (using service role key)
and writes a .xlsx Muster Roll matching the Zenith template format.

Usage:
    python export_attendance_sheet.py --community-id <uuid> --month 3 --year 2026

Requires: SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY in .env (or environment)
"""

import argparse
import calendar
import json
import os
import sys
from datetime import date

try:
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(__file__), '..', '.env'))
except ImportError:
    pass  # dotenv optional if env vars are set directly

try:
    from supabase import create_client
except ImportError:
    print("ERROR: supabase-py not installed. Run: pip install -r tools/requirements.txt")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install -r tools/requirements.txt")
    sys.exit(1)

DAY_NAMES = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
SHORT_DAY = ['M', 'T', 'W', 'Th', 'F', 'Sa', 'Su']

# Colour palette matching CoachOps design system
COLOR_HEADER = '0a0a0a'      # near black
COLOR_ACCENT = 'e8ff47'      # neon yellow (primary)
COLOR_PH = '4ade80'          # green
COLOR_ABSENT = 'f87171'      # red
COLOR_SUB = 'fbbf24'         # amber
COLOR_WKOFF = '252a35'       # dark grey


def make_client():
    url = os.environ.get('VITE_SUPABASE_URL') or os.environ.get('SUPABASE_URL')
    key = os.environ.get('SUPABASE_SERVICE_ROLE_KEY')
    if not url or not key:
        raise EnvironmentError(
            "Missing SUPABASE_URL/VITE_SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY."
            " Set them in .env or environment."
        )
    return create_client(url, key)


def fetch_data(client, community_id: str, month: int, year: int):
    month_start = f"{year}-{month:02d}-01"
    total_days = calendar.monthrange(year, month)[1]
    month_end = f"{year}-{month:02d}-{total_days:02d}"

    # Community info
    comm = client.table('communities').select('name').eq('id', community_id).single().execute()
    community_name = comm.data['name']

    # Sports for this community
    sports = client.table('sports').select('*').eq('community_id', community_id).execute().data

    # Coaches and assignments for this community's sports
    sport_ids = [s['id'] for s in sports]
    assignments = client.table('coach_sport_assignments') \
        .select('coach_id, sport_id, monthly_salary, coaches(id, name)') \
        .in_('sport_id', sport_ids) \
        .execute().data

    # Attendance records for the month
    attendance_records = client.table('monthly_attendance') \
        .select('*') \
        .eq('community_id', community_id) \
        .eq('month', month) \
        .eq('year', year) \
        .execute().data
    att_index = {(r['coach_id'], r['sport_id']): r for r in attendance_records}

    # Paid holidays
    ph_data = client.table('paid_holidays') \
        .select('date, holiday_name') \
        .eq('community_id', community_id) \
        .gte('date', month_start) \
        .lte('date', month_end) \
        .execute().data
    ph_days = {date.fromisoformat(p['date']).day: p['holiday_name'] for p in ph_data}

    return community_name, sports, assignments, att_index, ph_days, total_days


def export_sheet(community_id: str, month: int, year: int, output_path: str | None = None):
    client = make_client()
    community_name, sports, assignments, att_index, ph_days, total_days = fetch_data(
        client, community_id, month, year
    )

    month_name = date(year, month, 1).strftime('%B %Y')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    op_set_for = {s['id']: set(s['operating_days']) for s in sports}
    sports_by_id = {s['id']: s for s in sports}

    for sport in sports:
        ws = wb.create_sheet(title=sport['sport_name'][:31])
        op_set = op_set_for[sport['id']]
        sport_coaches = [a for a in assignments if a['sport_id'] == sport['id']]

        _write_muster_sheet(
            ws=ws,
            community_name=community_name,
            sport=sport,
            coaches=sport_coaches,
            att_index=att_index,
            ph_days=ph_days,
            total_days=total_days,
            month=month,
            year=year,
            op_set=op_set,
        )

    if not output_path:
        safe_name = community_name.replace(' ', '_')
        month_str = date(year, month, 1).strftime('%b_%Y')
        output_path = f"CoachOps_Attendance_{safe_name}_{month_str}.xlsx"

    wb.save(output_path)
    print(f"Saved: {output_path}")
    return output_path


def _thin_border():
    side = Side(style='thin', color='252a35')
    return Border(left=side, right=side, top=side, bottom=side)


def _write_muster_sheet(ws, community_name, sport, coaches, att_index, ph_days, total_days,
                         month, year, op_set):
    """Write one sport's muster roll to the worksheet."""

    # ── Title rows ────────────────────────────────────────────────────────────
    ws.append([f"COACHOPS — MUSTER ROLL"])
    ws.append([f"{community_name} | {sport['sport_name']}"])
    ws.append([f"{date(year, month, 1).strftime('%B %Y')} | Weekly Off: {sport['weekly_off_day']}"])
    ws.append([])  # blank row

    # ── Header row: S.No | Coach | 1 2 3 … 31 | WD | PH | A | SUB | PaidD | Salary ──
    header = ['S.No', 'Coach Name']
    day_to_col = {}  # day_num → col index in header list
    for d in range(1, total_days + 1):
        day_to_col[d] = len(header)
        header.append(d)
    summary_cols = ['WD', 'PH', 'A', 'SUB', 'Paid Days', 'Salary']
    header.extend(summary_cols)
    ws.append(header)

    header_row = ws.max_row
    # Style header
    for col_idx, val in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = Font(name='DM Sans', bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor=COLOR_HEADER)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Mark PH columns in header
    for d, holiday_name in ph_days.items():
        col_idx = day_to_col[d] + 1  # 1-indexed
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = PatternFill('solid', fgColor=COLOR_PH)
        cell.comment = None  # openpyxl comments require extra package

    # ── Data rows ──────────────────────────────────────────────────────────────
    for idx, coach_assign in enumerate(coaches, start=1):
        coach = coach_assign['coaches']
        sport_id = coach_assign['sport_id']
        monthly_salary = float(coach_assign['monthly_salary'])
        rec = att_index.get((coach['id'], sport_id), {})
        att_data = rec.get('attendance_data') or {}
        total_wd = rec.get('total_working_days', 0)
        ph_count = rec.get('paid_holidays', 0)
        absences = rec.get('days_absent', 0)
        subs = rec.get('days_substitute', 0)
        paid_days = rec.get('total_paid_days', total_wd)
        calc_salary = paid_days / total_wd * monthly_salary if total_wd > 0 else 0

        row_data = [idx, coach['name']]
        # Day cells
        for d in range(1, total_days + 1):
            day_date = date(year, month, d)
            day_name = DAY_NAMES[day_date.weekday()]
            if day_name == sport['weekly_off_day']:
                row_data.append('WO')
            elif day_name not in op_set:
                row_data.append('')
            elif d in ph_days:
                row_data.append('PH')
            else:
                code = att_data.get(str(d), 'P')
                row_data.append(code)

        row_data.extend([total_wd, ph_count, absences, subs, paid_days, f'=ROUND({paid_days}/{total_wd if total_wd else 1}*{monthly_salary},2)'])
        ws.append(row_data)

        data_row = ws.max_row
        for col_idx in range(1, len(header) + 1):
            cell = ws.cell(row=data_row, column=col_idx)
            cell.border = _thin_border()
            cell.alignment = Alignment(horizontal='center')
            val = cell.value
            if val == 'A':
                cell.fill = PatternFill('solid', fgColor=COLOR_ABSENT)
                cell.font = Font(bold=True, color='FFFFFF')
            elif val == 'SUB':
                cell.fill = PatternFill('solid', fgColor=COLOR_SUB)
                cell.font = Font(bold=True)
            elif val == 'PH':
                cell.fill = PatternFill('solid', fgColor=COLOR_PH)
                cell.font = Font(bold=True)
            elif val == 'WO':
                cell.fill = PatternFill('solid', fgColor=COLOR_WKOFF)
                cell.font = Font(color='888888')

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 6   # S.No
    ws.column_dimensions['B'].width = 22  # Coach name
    for d in range(1, total_days + 1):
        col_letter = get_column_letter(day_to_col[d] + 1)
        ws.column_dimensions[col_letter].width = 4.5
    # Summary cols
    for i, col_name in enumerate(summary_cols, start=day_to_col[total_days] + 2):
        ws.column_dimensions[get_column_letter(i)].width = 10


def main():
    parser = argparse.ArgumentParser(description='Export CoachOps attendance sheet to Excel')
    parser.add_argument('--community-id', required=True, help='Community UUID')
    parser.add_argument('--month', type=int, required=True, help='Month 1-12')
    parser.add_argument('--year', type=int, required=True, help='Year e.g. 2026')
    parser.add_argument('--output', help='Output file path (optional)')
    args = parser.parse_args()

    export_sheet(args.community_id, args.month, args.year, args.output)


if __name__ == '__main__':
    main()
